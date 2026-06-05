"""Weekly job timesheet export — Excel template fill + portrait PDF."""

from __future__ import annotations

import base64
import html
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore[misc, assignment]
    load_workbook = None  # type: ignore[misc, assignment]

try:
    from app.branding import get_header_logo_path
    from app.services.job_weekly_timesheets import monday_of_week
    from app.services.weekly_job_timesheet_service import (
        TimesheetLine,
        WeeklyJobTimesheetData,
        TIMESHEET_LABOR_MIN_ROWS,
        TIMESHEET_MATERIAL_MIN_ROWS,
        _day_labels,
        _fmt_hours,
        _fmt_money,
        _pad_labor_rows,
        _pad_material_rows,
        _parse_date,
        _split_materials,
        load_timesheet_data,
        render_timesheet_html,
    )
except ImportError:
    from branding import get_header_logo_path  # type: ignore
    from services.job_weekly_timesheets import monday_of_week  # type: ignore
    from services.weekly_job_timesheet_service import (  # type: ignore
        TimesheetLine,
        WeeklyJobTimesheetData,
        TIMESHEET_LABOR_MIN_ROWS,
        TIMESHEET_MATERIAL_MIN_ROWS,
        _day_labels,
        _fmt_hours,
        _fmt_money,
        _pad_labor_rows,
        _pad_material_rows,
        _parse_date,
        _split_materials,
        load_timesheet_data,
        render_timesheet_html,
    )

_ROOT = Path(__file__).resolve().parents[2]
_TEMPLATE_PATH = _ROOT / "templates" / "TIMESHEET WEEKLY.xlsx"
_HTML_TEMPLATE = Path(__file__).resolve().parents[1] / "templates" / "weekly_job_timesheet.html"

# ---------------------------------------------------------------------------
# TIMESHEET WEEKLY.xlsx cell mapping (portrait letter, 13 columns A–M)
# ---------------------------------------------------------------------------
# Row 5–9 (col A label, col B value): JOB #, CLIENT, JOB NAME, P.O. #, DATE/WEEK
# Row 11: labor column headers (A11:M11)
# Rows 12–23: labor lines (12 rows)
# Row 25+: materials; row 36+: work performed; row 41+: approval
# ---------------------------------------------------------------------------
_CELL = {
    "job_number": "B5",
    "client_name": "B6",
    "job_name": "B7",
    "po_number": "B8",
    "sheet_date": "B9",
    "work_performed": "A37",
    "approved_by": "A42",
    "signed_date": "J42",
}
_FIRST_LABOR_ROW = 12
_LABOR_ROW_COUNT = 15
_FIRST_MATERIAL_ROW = 27
_MATERIAL_ROW_COUNT = 10
_PDF_ROW_HEIGHT = 0.26 * inch
_PDF_HEAD_HEIGHT = 0.30 * inch
_PDF_HEADER_HEIGHT = 0.75 * inch
_PDF_APPROVAL_HEIGHT = 1.05 * inch


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left_wrap() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def ensure_excel_template() -> Path:
    """Create templates/TIMESHEET WEEKLY.xlsx if missing."""
    _TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if _TEMPLATE_PATH.is_file():
        return _TEMPLATE_PATH
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel export. Install openpyxl>=3.1.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Timesheet"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    widths = {"A": 22, "B": 10, "C": 5.5, "D": 5.5, "E": 5.5, "F": 5.5, "G": 5.5, "H": 5.5, "I": 5.5, "J": 4.5, "K": 4.5, "L": 4.5, "M": 5.5}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    header_fill = PatternFill("solid", fgColor="E8EEF7")
    section_fill = PatternFill("solid", fgColor="DBEAFE")
    bold = Font(bold=True, size=9)
    normal = Font(size=9)

    ws.merge_cells("A1:D4")
    ws["A1"].alignment = _center()
    ws["E1"] = "WEEKLY TIMESHEET"
    ws["E1"].font = Font(bold=True, size=16)
    ws["E1"].alignment = _center()
    ws.merge_cells("E1:M1")
    ws["E2"] = "INDUSTRIAL PLANT SOLUTIONS, LLC"
    ws["E2"].font = Font(bold=True, size=9)
    ws["E2"].alignment = _center()
    ws.merge_cells("E2:M2")

    meta = [("JOB #", ""), ("CLIENT", ""), ("JOB NAME", ""), ("P.O. #", ""), ("DATE / WEEK", "")]
    for i, (lbl, _) in enumerate(meta, start=5):
        ws[f"A{i}"] = lbl
        ws[f"A{i}"].font = bold
        ws[f"B{i}"].font = normal
        ws[f"A{i}"].border = _thin_border()
        ws[f"B{i}"].border = _thin_border()
        ws.merge_cells(f"B{i}:D{i}")

    logo_path = get_header_logo_path()
    if logo_path and logo_path.is_file():
        try:
            img = XLImage(str(logo_path))
            img.width = 90
            img.height = 48
            ws.add_image(img, "A1")
        except Exception:
            ws["A1"] = "IPS"

    labor_headers = [
        "EMPLOYEE / EQUIPMENT RENTAL",
        "CLASS",
        "MON",
        "TUE",
        "WED",
        "THU",
        "FRI",
        "SAT",
        "SUN",
        "ST",
        "OT",
        "TOTAL",
    ]
    for col_idx, hdr in enumerate(labor_headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=hdr)
        cell.font = bold
        cell.fill = header_fill
        cell.border = _thin_border()
        cell.alignment = _center()

    for r in range(_FIRST_LABOR_ROW, _FIRST_LABOR_ROW + _LABOR_ROW_COUNT):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.border = _thin_border()
            cell.alignment = _center() if c > 2 else _left_wrap()

    ws.merge_cells("A25:C25")
    ws["A25"] = "MATERIALS / EXPENSES"
    ws["A25"].fill = section_fill
    ws["A25"].font = bold
    ws["A25"].alignment = _center()
    ws.merge_cells("E25:G25")
    ws["E25"] = "MATERIALS / EXPENSES (CONT.)"
    ws["E25"].fill = section_fill
    ws["E25"].font = bold
    ws["E25"].alignment = _center()

    for lbl, col in (("DESCRIPTION", "A"), ("QTY", "B"), ("COST", "C"), ("DESCRIPTION", "E"), ("QTY", "F"), ("COST", "G")):
        cell = ws[f"{col}26"]
        cell.value = lbl
        cell.font = bold
        cell.fill = header_fill
        cell.border = _thin_border()
        cell.alignment = _center()

    for r in range(_FIRST_MATERIAL_ROW, _FIRST_MATERIAL_ROW + _MATERIAL_ROW_COUNT):
        for col in ("A", "B", "C", "E", "F", "G"):
            ws[f"{col}{r}"].border = _thin_border()

    ws.merge_cells("A36:M36")
    ws["A36"] = "WORK PERFORMED"
    ws["A36"].font = bold
    ws["A36"].fill = PatternFill("solid", fgColor="F1F5F9")
    ws.merge_cells("A37:M39")
    ws["A37"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A41"] = "APPROVED BY"
    ws["F41"] = "SIGNATURE"
    ws["J41"] = "DATE SIGNED"
    for addr in ("A41", "F41", "J41"):
        ws[addr].font = bold

    wb.save(_TEMPLATE_PATH)
    return _TEMPLATE_PATH


def build_timesheet_template_context(timesheet_id: str) -> dict[str, Any]:
    """Load timesheet + computed export context."""
    data = load_timesheet_data(timesheet_id)
    if not data:
        raise ValueError("Timesheet not found.")
    ws = _parse_date(data.week_start) or date.today()
    ws = monday_of_week(ws)
    labor = _pad_labor_rows([ln for ln in data.labor_lines if ln.line_type in {"labor", "equipment"}], min_rows=_LABOR_ROW_COUNT)
    mats_left, mats_right = _split_materials(data.material_lines)
    mat_left = _pad_material_rows(mats_left, _MATERIAL_ROW_COUNT)
    mat_right = _pad_material_rows(mats_right, _MATERIAL_ROW_COUNT)
    week_label = f"{data.week_start} – {data.week_end}" if data.week_end else data.sheet_date
    return {
        "data": data,
        "week_start": ws,
        "day_labels": _day_labels(ws),
        "labor_lines": labor,
        "materials_left": mat_left,
        "materials_right": mat_right,
        "week_label": week_label,
        "total_hours": round(sum(ln.total_hours for ln in labor), 2),
        "materials_total": round(sum(ln.cost for ln in data.material_lines), 2),
    }


def _set_cell(ws, addr: str, value: Any, *, align_center: bool = False) -> None:
    ws[addr] = value
    ws[addr].alignment = _center() if align_center else _left_wrap()
    ws[addr].font = Font(size=9)


def _fill_labor_row(ws, row: int, ln: Any) -> None:
    values = [
        ln.description,
        ln.class_name,
        ln.mon,
        ln.tue,
        ln.wed,
        ln.thu,
        ln.fri,
        ln.sat,
        ln.sun,
        ln.st_hours,
        ln.ot_hours + ln.dt_hours,
        ln.total_hours,
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx)
        if col_idx >= 3 and isinstance(val, (int, float)):
            cell.value = val if val else None
            cell.number_format = "0.00"
        else:
            cell.value = val or None
        cell.border = _thin_border()
        cell.alignment = _center() if col_idx >= 3 else _left_wrap()


def export_data_to_excel_bytes(data: WeeklyJobTimesheetData) -> bytes:
    """Excel export from unsaved preview data."""
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for Excel export.")
    ws_date = _parse_date(data.week_start) or date.today()
    ws_date = monday_of_week(ws_date)
    ctx = {
        "data": data,
        "week_start": ws_date,
        "day_labels": _day_labels(ws_date),
        "labor_lines": _pad_labor_rows(
            [ln for ln in data.labor_lines if ln.line_type in {"labor", "equipment"}],
            min_rows=_LABOR_ROW_COUNT,
        ),
        "materials_left": _pad_material_rows(_split_materials(data.material_lines)[0], _MATERIAL_ROW_COUNT),
        "materials_right": _pad_material_rows(_split_materials(data.material_lines)[1], _MATERIAL_ROW_COUNT),
        "week_label": f"{data.week_start} – {data.week_end}" if data.week_end else data.sheet_date,
    }
    return _fill_workbook_bytes(ctx)


def export_weekly_timesheet_to_excel(timesheet_id: str) -> bytes:
    """Fill templates/TIMESHEET WEEKLY.xlsx and return bytes."""
    ctx = build_timesheet_template_context(timesheet_id)
    return _fill_workbook_bytes(ctx)


def _fill_workbook_bytes(ctx: dict[str, Any]) -> bytes:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for Excel export.")
    data: WeeklyJobTimesheetData = ctx["data"]
    path = ensure_excel_template()
    wb = load_workbook(path)
    ws = wb.active

    _set_cell(ws, _CELL["job_number"], data.job_number)
    _set_cell(ws, _CELL["client_name"], data.client_name)
    _set_cell(ws, _CELL["job_name"], data.job_name)
    _set_cell(ws, _CELL["po_number"], data.po_number)
    _set_cell(ws, _CELL["sheet_date"], ctx["week_label"] or data.sheet_date)

    day_labels = ctx["day_labels"]
    for i, lbl in enumerate(day_labels):
        parts = lbl.replace("<br>", "\n").split("\n")
        ws.cell(row=11, column=3 + i, value=f"{parts[0]}\n{parts[-1]}" if len(parts) > 1 else parts[0])
        ws.cell(row=11, column=3 + i).alignment = _center()

    labor_lines = ctx["labor_lines"]
    for i in range(_LABOR_ROW_COUNT):
        row = _FIRST_LABOR_ROW + i
        ln = labor_lines[i] if i < len(labor_lines) else None
        if ln is None:
            for c in range(1, 14):
                ws.cell(row=row, column=c, value=None)
            continue
        _fill_labor_row(ws, row, ln)

    left = ctx["materials_left"]
    right = ctx["materials_right"]
    for i in range(_MATERIAL_ROW_COUNT):
        row = _FIRST_MATERIAL_ROW + i
        l = left[i] if i < len(left) else None
        r = right[i] if i < len(right) else None
        if l:
            ws[f"A{row}"] = l.description
            ws[f"B{row}"] = l.qty or None
            ws[f"C{row}"] = l.cost or None
        if r:
            ws[f"E{row}"] = r.description
            ws[f"F{row}"] = r.qty or None
            ws[f"G{row}"] = r.cost or None

    ws[_CELL["work_performed"]] = data.work_performed
    ws[_CELL["approved_by"]] = data.approved_by
    ws[_CELL["signed_date"]] = (data.signed_at or "")[:10]

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_weekly_timesheet_to_pdf(timesheet_id: str) -> bytes:
    """Portrait letter PDF matching the weekly timesheet layout."""
    ctx = build_timesheet_template_context(timesheet_id)
    data: WeeklyJobTimesheetData = ctx["data"]
    return _build_portrait_pdf(data, week_start=ctx["week_start"])


def export_data_to_pdf(data: WeeklyJobTimesheetData) -> bytes:
    """PDF from in-memory timesheet data (preview/download before save)."""
    ws = _parse_date(data.week_start) or date.today()
    return _build_portrait_pdf(data, week_start=monday_of_week(ws))


def _build_portrait_pdf(data: WeeklyJobTimesheetData, *, week_start: date) -> bytes:
    labels = _day_labels(week_start)
    buf = BytesIO()
    margin = 0.25 * inch
    page_w, page_h = letter
    content_w = page_w - (2 * margin)
    content_h = page_h - (2 * margin)
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []

    logo_path = get_header_logo_path()
    title_row: list[Any]
    if logo_path and logo_path.is_file():
        try:
            title_row = [
                RLImage(str(logo_path), width=0.95 * inch, height=0.55 * inch),
                Paragraph(
                    "<b>WEEKLY TIMESHEET</b><br/><font size=8>INDUSTRIAL PLANT SOLUTIONS, LLC</font>",
                    styles["Normal"],
                ),
            ]
        except Exception:
            title_row = [Paragraph("<b>IPS</b>", styles["Normal"]), Paragraph("<b>WEEKLY TIMESHEET</b>", styles["Normal"])]
    else:
        title_row = [
            Paragraph("<b>IPS</b>", styles["Normal"]),
            Paragraph(
                "<b>WEEKLY TIMESHEET</b><br/><font size=8>INDUSTRIAL PLANT SOLUTIONS, LLC</font>",
                styles["Normal"],
            ),
        ]

    meta = [
        ["JOB #", data.job_number],
        ["CLIENT", data.client_name],
        ["JOB NAME", data.job_name],
        ["P.O. #", data.po_number],
        ["DATE", f"{data.week_start} – {data.week_end}"],
    ]
    meta_tbl = Table(meta, colWidths=[0.72 * inch, 1.75 * inch])
    meta_tbl.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ]
        )
    )
    header_tbl = Table(
        [[title_row[0], title_row[1], meta_tbl]],
        colWidths=[0.95 * inch, content_w - 0.95 * inch - 2.47 * inch, 2.47 * inch],
        rowHeights=[_PDF_HEADER_HEIGHT],
    )
    header_tbl.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    story.append(header_tbl)

    labor = _pad_labor_rows(
        [ln for ln in data.labor_lines if ln.line_type in {"labor", "equipment"}],
        min_rows=TIMESHEET_LABOR_MIN_ROWS,
    )
    header = ["Employee / Equipment", "Class"] + [lbl.replace("<br>", "\n") for lbl in labels] + ["ST", "OT", "Total"]
    grid: list[list[str]] = [header]
    for ln in labor:
        grid.append(
            [
                ln.description,
                ln.class_name,
                _fmt_hours(ln.mon),
                _fmt_hours(ln.tue),
                _fmt_hours(ln.wed),
                _fmt_hours(ln.thu),
                _fmt_hours(ln.fri),
                _fmt_hours(ln.sat),
                _fmt_hours(ln.sun),
                _fmt_hours(ln.st_hours),
                _fmt_hours(ln.ot_hours + ln.dt_hours),
                _fmt_hours(ln.total_hours),
            ]
        )
    col_w = [
        content_w * 0.20,
        content_w * 0.08,
        *[content_w * 0.07] * 7,
        content_w * 0.05,
        content_w * 0.05,
        content_w * 0.08,
    ]
    labor_heights = [_PDF_HEAD_HEIGHT] + [_PDF_ROW_HEIGHT] * len(labor)
    labor_tbl = Table(grid, colWidths=col_w, rowHeights=labor_heights, repeatRows=1)
    labor_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eef7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
                ("ALIGN", (2, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(labor_tbl)

    mats_left, mats_right = _split_materials(data.material_lines)
    mats_left = _pad_material_rows(mats_left, TIMESHEET_MATERIAL_MIN_ROWS)
    mats_right = _pad_material_rows(mats_right, TIMESHEET_MATERIAL_MIN_ROWS)
    mat_grid = [["Materials / Expenses", "Qty", "Cost", "Materials / Expenses (cont.)", "Qty", "Cost"]]
    empty_mat = TimesheetLine(line_type="material")
    for i in range(max(len(mats_left), len(mats_right))):
        l = mats_left[i] if i < len(mats_left) else empty_mat
        r = mats_right[i] if i < len(mats_right) else empty_mat
        mat_grid.append(
            [
                l.description,
                _fmt_hours(l.qty),
                _fmt_money(l.cost),
                r.description,
                _fmt_hours(r.qty),
                _fmt_money(r.cost),
            ]
        )
    half_w = content_w / 2.0
    mat_row_count = max(len(mats_left), len(mats_right))
    mat_heights = [_PDF_HEAD_HEIGHT] + [_PDF_ROW_HEIGHT] * mat_row_count
    mat_tbl = Table(
        mat_grid,
        colWidths=[half_w * 0.72, half_w * 0.14, half_w * 0.14, half_w * 0.72, half_w * 0.14, half_w * 0.14],
        rowHeights=mat_heights,
    )
    mat_tbl.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    story.append(mat_tbl)

    work_h = (
        content_h
        - _PDF_HEADER_HEIGHT
        - sum(labor_heights)
        - sum(mat_heights)
        - _PDF_APPROVAL_HEIGHT
    )
    work_h = max(work_h, 1.0 * inch)
    work_label_h = 0.24 * inch

    work_body = html.escape(data.work_performed or " ").replace("\n", "<br/>") or "&nbsp;"
    work_tbl = Table(
        [[Paragraph("<b>WORK PERFORMED</b>", styles["Heading4"])], [Paragraph(work_body, styles["BodyText"])]],
        colWidths=[content_w],
        rowHeights=[work_label_h, max(work_h - work_label_h, 0.75 * inch)],
    )
    work_tbl.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.35, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(work_tbl)

    approval_row = [
        Paragraph(f"<b>APPROVED BY</b><br/>{html.escape(data.approved_by or ' ')}", styles["Normal"]),
        Paragraph("<b>SIGNATURE</b>", styles["Normal"]),
        Paragraph(f"<b>DATE SIGNED</b><br/>{html.escape(data.signed_at[:10] if data.signed_at else ' ')}", styles["Normal"]),
    ]
    sig = str(data.signature_data or "").strip()
    if sig.startswith("data:image"):
        try:
            b64 = sig.split(",", 1)[1]
            approval_row[1] = RLImage(BytesIO(base64.b64decode(b64)), width=2.0 * inch, height=0.65 * inch)
        except Exception:
            approval_row[1] = Paragraph("<b>SIGNATURE</b>", styles["Normal"])
    appr_tbl = Table([approval_row], colWidths=[content_w / 3.0] * 3, rowHeights=[_PDF_APPROVAL_HEIGHT])
    appr_tbl.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.35, colors.black),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(appr_tbl)

    doc.build(story)
    return buf.getvalue()


def render_weekly_job_timesheet_html(timesheet_id: str) -> str:
    """HTML preview using weekly_job_timesheet.html when present."""
    data = load_timesheet_data(timesheet_id)
    if not data:
        raise ValueError("Timesheet not found.")
    ws = _parse_date(data.week_start) or date.today()
    if _HTML_TEMPLATE.is_file():
        return render_timesheet_html(data, week_start=monday_of_week(ws))
    return render_timesheet_html(data, week_start=monday_of_week(ws))
